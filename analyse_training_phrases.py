"""
Training Phrase Deep Analysis
==============================
Input : CSV file with columns (utterance, intent) — produced by phrase_extraction.py
Output: Excel report with 6 analytical sheets

Sheets
------
1. Summary               - per-intent phrase count, word count stats, flags
2. Token Overlap Matrix  - pairwise Jaccard similarity between every intent pair
3. High Overlap Pairs    - ranked list of intent pairs with overlap >= threshold
4. Shared Tokens Detail  - every shared token + which intents use it
5. Intent Word Stats     - full word-count distribution per intent
6. Action Items          - prioritised remediation tasks ready to action
"""

import csv
import re
import math
import itertools
import os
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from dotenv import load_dotenv

load_dotenv()

# ── Config — all values loaded from .env (see .env.example) ───────────────────
INPUT_CSV         = os.getenv("INPUT_CSV",  "training_phrases.csv")
OUTPUT_XLSX       = os.getenv("OUTPUT_XLSX", "training_phrase_analysis.xlsx")
OVERLAP_THRESHOLD = float(os.getenv("OVERLAP_THRESHOLD", 0.30))
LOW_PHRASE_MIN    = int(os.getenv("LOW_PHRASE_MIN", 5))
HIGH_WORD_MAX     = 15     # avg word count above this → flagged as too long
LOW_WORD_MIN      = 2      # avg word count below this → flagged as too short
TOP_N_SHARED      = 50     # top N most-shared tokens in the shared-token sheet

# Stopwords to exclude from token analysis (noise reduction)
STOPWORDS = {
    "i","my","me","a","an","the","to","for","of","in","on","at","is","it",
    "do","can","how","what","when","where","why","will","would","could",
    "should","please","want","need","get","have","has","be","with","about",
    "and","or","not","no","your","you","that","this","are","was","were",
    "from","if","so","up","by","as","but","its","than","into","there",
    "their","them","they","we","us","our","he","she","his","her","which",
    "been","being","also","just","more","any","all","some","much","many",
    "then","out","had","did","does","am","per","re","let","go","tell",
}

# ── Palette ────────────────────────────────────────────────────────────────────
C_HEADER_DARK  = "1F3864"
C_HEADER_MID   = "2E75B6"
C_HEADER_LIGHT = "BDD7EE"
C_ACCENT       = "ED7D31"
C_WARN         = "FFD966"
C_DANGER       = "FF6B6B"
C_OK           = "70AD47"
C_WHITE        = "FFFFFF"
C_LIGHT_GREY   = "F2F2F2"
C_STRIPE       = "DEEAF1"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# ── Tokeniser ─────────────────────────────────────────────────────────────────
def tokenise(text):
    """Lowercase alphanumeric tokens, stopwords removed."""
    tokens = re.findall(r"[a-z0-9]+", text.lower())
    return [t for t in tokens if t not in STOPWORDS and len(t) > 1]

def word_count(text):
    return len(text.strip().split())

# ── Load CSV ───────────────────────────────────────────────────────────────────
def load_csv(path):
    intent_phrases = defaultdict(list)
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            utt = row["utterance"].strip()
            intent = row["intent"].strip()
            if utt and intent:
                intent_phrases[intent].append(utt)
    return dict(intent_phrases)

# ── Core Analytics ─────────────────────────────────────────────────────────────
def compute_intent_stats(intent_phrases):
    stats = {}
    for intent, phrases in intent_phrases.items():
        wcs = [word_count(p) for p in phrases]
        tokens_all = []
        for p in phrases:
            tokens_all.extend(tokenise(p))
        stats[intent] = {
            "phrase_count": len(phrases),
            "avg_wc":        round(sum(wcs) / len(wcs), 2) if wcs else 0,
            "min_wc":        min(wcs) if wcs else 0,
            "max_wc":        max(wcs) if wcs else 0,
            "median_wc":     sorted(wcs)[len(wcs)//2] if wcs else 0,
            "token_set":     set(tokens_all),
            "token_counter": Counter(tokens_all),
            "wcs":           wcs,
        }
    return stats

def jaccard(set_a, set_b):
    inter = len(set_a & set_b)
    union = len(set_a | set_b)
    return round(inter / union, 4) if union else 0.0

def compute_pairwise_overlap(stats):
    intents = sorted(stats.keys())
    pairs = []
    for a, b in itertools.combinations(intents, 2):
        j = jaccard(stats[a]["token_set"], stats[b]["token_set"])
        shared = sorted(stats[a]["token_set"] & stats[b]["token_set"])
        pairs.append({
            "intent_a":     a,
            "intent_b":     b,
            "jaccard":      j,
            "shared_count": len(shared),
            "shared_tokens": shared,
        })
    pairs.sort(key=lambda x: x["jaccard"], reverse=True)
    return intents, pairs

def compute_shared_token_map(stats):
    """For each token, which intents use it and how many times."""
    token_to_intents = defaultdict(dict)
    for intent, s in stats.items():
        for token, count in s["token_counter"].items():
            token_to_intents[token][intent] = count
    # Filter to tokens shared across >= 2 intents
    shared = {t: d for t, d in token_to_intents.items() if len(d) >= 2}
    # Sort by number of intents using the token descending
    return dict(sorted(shared.items(), key=lambda x: len(x[1]), reverse=True))

# ── Sheet Helpers ──────────────────────────────────────────────────────────────
def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def write_header_row(ws, row_num, headers, bg=C_HEADER_DARK, fg=C_WHITE, height=18):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = font(bold=True, color=fg)
        cell.fill = fill(bg)
        cell.alignment = centre()
        cell.border = border_thin()
    ws.row_dimensions[row_num].height = height

def write_data_row(ws, row_num, values, stripe=False):
    bg = C_STRIPE if stripe else C_WHITE
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=v)
        cell.fill = fill(bg)
        cell.alignment = left_align(wrap=False)
        cell.border = border_thin()
        cell.font = Font(name="Arial", size=9, color="2C2C2C")

def add_title_block(ws, title, subtitle=""):
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = title
    t.font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    t.fill = fill(C_HEADER_DARK)
    t.alignment = centre()
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws.merge_cells("A2:H2")
        s = ws["A2"]
        s.value = subtitle
        s.font = Font(name="Arial", size=9, italic=True, color="595959")
        s.fill = fill(C_LIGHT_GREY)
        s.alignment = centre()
        ws.row_dimensions[2].height = 14

    return 3 if subtitle else 2  # next usable row

# ── Sheet 1: Summary ───────────────────────────────────────────────────────────
def build_summary(wb, stats, overlap_threshold):
    ws = wb.create_sheet("1. Summary")
    next_row = add_title_block(
        ws,
        "Training Phrase Analysis — Intent Summary",
        f"Flags: < {LOW_PHRASE_MIN} phrases | avg word count < {LOW_WORD_MIN} or > {HIGH_WORD_MAX} | high token overlap elsewhere"
    )

    headers = [
        "Intent", "Phrase Count", "Avg Word Count",
        "Min WC", "Max WC", "Median WC",
        "Unique Tokens", "Flag"
    ]
    write_header_row(ws, next_row, headers)
    next_row += 1

    intents_sorted = sorted(stats.keys(), key=lambda x: x.lower())
    for i, intent in enumerate(intents_sorted):
        s = stats[intent]
        flags = []
        if s["phrase_count"] < LOW_PHRASE_MIN:
            flags.append(f"Low phrases ({s['phrase_count']})")
        if s["avg_wc"] > HIGH_WORD_MAX:
            flags.append(f"Avg WC high ({s['avg_wc']})")
        if s["avg_wc"] < LOW_WORD_MIN:
            flags.append(f"Avg WC low ({s['avg_wc']})")

        flag_str = " | ".join(flags) if flags else "OK"
        row_vals = [
            intent,
            s["phrase_count"],
            s["avg_wc"],
            s["min_wc"],
            s["max_wc"],
            s["median_wc"],
            len(s["token_set"]),
            flag_str,
        ]
        write_data_row(ws, next_row, row_vals, stripe=(i % 2 == 0))

        # Colour the flag cell
        flag_cell = ws.cell(row=next_row, column=8)
        if flag_str == "OK":
            flag_cell.fill = fill("E2EFDA")
            flag_cell.font = Font(name="Arial", size=9, color="375623", bold=True)
        else:
            flag_cell.fill = fill("FFE699")
            flag_cell.font = Font(name="Arial", size=9, color="7F6000", bold=True)

        # Colour phrase count < LOW_PHRASE_MIN
        pc_cell = ws.cell(row=next_row, column=2)
        if s["phrase_count"] < LOW_PHRASE_MIN:
            pc_cell.fill = fill("FFD7D7")
            pc_cell.font = Font(name="Arial", size=9, color="C00000", bold=True)

        next_row += 1

    # Totals row
    ws.cell(row=next_row, column=1, value="TOTAL / AVERAGE").font = Font(
        name="Arial", bold=True, size=9)
    ws.cell(row=next_row, column=2, value=f'=SUM(B{next_row-len(intents_sorted)}:B{next_row-1})')
    ws.cell(row=next_row, column=3, value=f'=AVERAGE(C{next_row-len(intents_sorted)}:C{next_row-1})')
    for col in range(1, 9):
        c = ws.cell(row=next_row, column=col)
        c.fill = fill(C_HEADER_LIGHT)
        c.font = Font(name="Arial", bold=True, size=9, color="1F3864")
        c.border = border_thin()

    set_col_widths(ws, [38, 14, 16, 9, 9, 12, 15, 38])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{next_row - len(intents_sorted) - 1}:H{next_row - 1}"

# ── Sheet 2: Token Overlap Matrix ─────────────────────────────────────────────
def build_overlap_matrix(wb, intents, stats):
    ws = wb.create_sheet("2. Overlap Matrix")
    add_title_block(
        ws,
        "Token Overlap Matrix (Jaccard Similarity)",
        "Cell value = Jaccard similarity (shared tokens / union tokens). Higher = more overlap. Diagonal = 1.0"
    )

    # Cap matrix at 60 intents to stay readable
    display_intents = intents[:]
    offset_row = 4
    offset_col = 2

    # Column headers
    ws.cell(row=offset_row, column=1, value="Intent").fill = fill(C_HEADER_DARK)
    ws.cell(row=offset_row, column=1).font = font(bold=True)
    ws.cell(row=offset_row, column=1).alignment = centre()

    for j, intent in enumerate(display_intents):
        c = ws.cell(row=offset_row, column=offset_col + j, value=intent)
        c.font = Font(name="Arial", bold=True, size=7, color=C_WHITE)
        c.fill = fill(C_HEADER_MID)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                text_rotation=45, wrap_text=False)
        ws.column_dimensions[get_column_letter(offset_col + j)].width = 4.5
    ws.row_dimensions[offset_row].height = 90

    # Row headers + values
    for i, intent_a in enumerate(display_intents):
        r = offset_row + 1 + i
        row_label = ws.cell(row=r, column=1, value=intent_a)
        row_label.font = Font(name="Arial", bold=True, size=8, color="1F3864")
        row_label.fill = fill(C_HEADER_LIGHT)
        row_label.alignment = left_align()
        ws.column_dimensions["A"].width = 36

        for j, intent_b in enumerate(display_intents):
            c = ws.cell(row=r, column=offset_col + j)
            if i == j:
                c.value = 1.0
                c.fill = fill("1F3864")
                c.font = Font(name="Arial", size=7, color=C_WHITE, bold=True)
            else:
                j_val = jaccard(stats[intent_a]["token_set"],
                                stats[intent_b]["token_set"])
                c.value = j_val
                c.font = Font(name="Arial", size=7)
                # Manual colour gradient: white → amber → red
                if j_val >= 0.6:
                    c.fill = fill("FF4444")
                    c.font = Font(name="Arial", size=7, color=C_WHITE, bold=True)
                elif j_val >= 0.4:
                    c.fill = fill("FF9933")
                    c.font = Font(name="Arial", size=7, color=C_WHITE)
                elif j_val >= 0.25:
                    c.fill = fill("FFD966")
                elif j_val >= 0.1:
                    c.fill = fill("EBF3FB")
                else:
                    c.fill = fill(C_WHITE)
            c.number_format = "0.00"
            c.alignment = centre()

        ws.row_dimensions[r].height = 13

    if len(intents) > 60:
        note_row = offset_row + len(display_intents) + 2
        ws.cell(row=note_row, column=1,
                value=f"⚠ Matrix capped at 60 intents for readability. "
                      f"{len(intents) - 60} intents omitted. See sheet '3. High Overlap Pairs' for full list.")
        ws.cell(row=note_row, column=1).font = Font(
            name="Arial", size=9, italic=True, color="C00000")

# ── Sheet 3: High Overlap Pairs ────────────────────────────────────────────────
def build_high_overlap(wb, pairs, threshold):
    ws = wb.create_sheet("3. High Overlap Pairs")
    next_row = add_title_block(
        ws,
        f"High Token Overlap Pairs  (Jaccard ≥ {threshold})",
        "These intent pairs share enough vocabulary that the NLU may confuse them. Review and diversify."
    )

    headers = ["#", "Intent A", "Intent B", "Jaccard Score",
               "Shared Tokens", "Shared Token List (sample)"]
    write_header_row(ws, next_row, headers)
    next_row += 1

    high = [p for p in pairs if p["jaccard"] >= threshold]
    if not high:
        ws.cell(row=next_row, column=1,
                value=f"✅  No intent pairs exceed the {threshold} Jaccard threshold. Vocabulary looks well-differentiated.")
        ws.cell(row=next_row, column=1).font = Font(
            name="Arial", size=10, color="375623", bold=True)
        return

    for i, p in enumerate(high):
        sample = ", ".join(p["shared_tokens"][:20])
        if len(p["shared_tokens"]) > 20:
            sample += f"  (+{len(p['shared_tokens'])-20} more)"
        vals = [i+1, p["intent_a"], p["intent_b"], p["jaccard"],
                p["shared_count"], sample]
        write_data_row(ws, next_row, vals, stripe=(i % 2 == 0))

        # Colour score cell
        score_cell = ws.cell(row=next_row, column=4)
        score_cell.number_format = "0.00%"
        if p["jaccard"] >= 0.6:
            score_cell.fill = fill("FF4444")
            score_cell.font = Font(name="Arial", size=9, color=C_WHITE, bold=True)
        elif p["jaccard"] >= 0.4:
            score_cell.fill = fill("FF9933")
            score_cell.font = Font(name="Arial", size=9, color=C_WHITE, bold=True)
        else:
            score_cell.fill = fill("FFD966")
            score_cell.font = Font(name="Arial", size=9, color="7F6000", bold=True)

        next_row += 1

    set_col_widths(ws, [5, 36, 36, 14, 14, 60])
    ws.freeze_panes = "A4"

# ── Sheet 4: Shared Tokens Detail ─────────────────────────────────────────────
def build_shared_tokens(wb, shared_token_map, top_n):
    ws = wb.create_sheet("4. Shared Token Detail")
    next_row = add_title_block(
        ws,
        f"Top {top_n} Tokens Shared Across Multiple Intents",
        "Tokens appearing in 2+ intents — high-frequency shared tokens weaken NLU discrimination."
    )

    headers = ["Rank", "Token", "# Intents Using It",
               "Total Occurrences", "Intents (alphabetical)"]
    write_header_row(ws, next_row, headers)
    next_row += 1

    top_tokens = list(shared_token_map.items())[:top_n]
    for rank, (token, intent_dict) in enumerate(top_tokens, 1):
        total_occ = sum(intent_dict.values())
        intent_list = ", ".join(sorted(intent_dict.keys()))
        vals = [rank, token, len(intent_dict), total_occ, intent_list]
        write_data_row(ws, next_row, vals, stripe=(rank % 2 == 0))

        # Colour the intent count cell
        count_cell = ws.cell(row=next_row, column=3)
        if len(intent_dict) >= 10:
            count_cell.fill = fill("FF4444")
            count_cell.font = Font(name="Arial", size=9, color=C_WHITE, bold=True)
        elif len(intent_dict) >= 5:
            count_cell.fill = fill("FFD966")
            count_cell.font = Font(name="Arial", size=9, color="7F6000", bold=True)

        next_row += 1

    set_col_widths(ws, [6, 18, 18, 18, 90])
    ws.freeze_panes = "A4"

# ── Sheet 5: Word Count Distribution ──────────────────────────────────────────
def build_word_stats(wb, stats, intent_phrases):
    ws = wb.create_sheet("5. Word Count Distribution")
    next_row = add_title_block(
        ws,
        "Word Count Distribution per Intent",
        "Breakdown of how many phrases fall into each word-count bucket."
    )

    buckets = ["1-2 wds", "3-5 wds", "6-10 wds", "11-15 wds", "16+ wds"]
    headers = ["Intent", "Total Phrases"] + buckets + ["Avg WC", "Assessment"]
    write_header_row(ws, next_row, headers)
    next_row += 1

    intents_sorted = sorted(stats.keys(), key=lambda x: x.lower())
    for i, intent in enumerate(intents_sorted):
        s = stats[intent]
        wcs = s["wcs"]
        b = [0] * 5
        for w in wcs:
            if w <= 2:   b[0] += 1
            elif w <= 5: b[1] += 1
            elif w <= 10:b[2] += 1
            elif w <= 15:b[3] += 1
            else:        b[4] += 1

        avg = s["avg_wc"]
        if avg < LOW_WORD_MIN:
            assessment = "⚠ Too short — add more descriptive phrases"
        elif avg > HIGH_WORD_MAX:
            assessment = "⚠ Too long — add shorter, natural utterances"
        else:
            assessment = "✅ Good range"

        vals = [intent, s["phrase_count"]] + b + [avg, assessment]
        write_data_row(ws, next_row, vals, stripe=(i % 2 == 0))

        assess_cell = ws.cell(row=next_row, column=len(headers))
        if "⚠" in assessment:
            assess_cell.fill = fill("FFE699")
            assess_cell.font = Font(name="Arial", size=9, color="7F6000", bold=True)
        else:
            assess_cell.fill = fill("E2EFDA")
            assess_cell.font = Font(name="Arial", size=9, color="375623")

        avg_cell = ws.cell(row=next_row, column=len(headers)-1)
        avg_cell.number_format = "0.0"

        next_row += 1

    set_col_widths(ws, [38, 14, 10, 10, 10, 10, 10, 10, 42])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{next_row - 1}"

# ── Sheet 6: Action Items ──────────────────────────────────────────────────────
def build_action_items(wb, stats, pairs, shared_token_map, threshold):
    ws = wb.create_sheet("6. Action Items")
    next_row = add_title_block(
        ws,
        "Prioritised Action Items",
        "Ready-to-use remediation tasks derived from the analysis. Tackle HIGH priority first."
    )

    headers = ["Priority", "Category", "Intent / Pair", "Issue", "Recommended Action"]
    write_header_row(ws, next_row, headers, bg=C_ACCENT)
    next_row += 1

    actions = []

    # LOW PHRASE COUNT
    for intent, s in sorted(stats.items()):
        if s["phrase_count"] < LOW_PHRASE_MIN:
            actions.append((
                "HIGH",
                "Insufficient Training Data",
                intent,
                f"Only {s['phrase_count']} training phrase(s) — NLU model will underfit this intent.",
                f"Add at least {LOW_PHRASE_MIN - s['phrase_count']} more varied phrases covering different ways users express this intent."
            ))

    # HIGH OVERLAP PAIRS
    critical = [p for p in pairs if p["jaccard"] >= 0.6]
    high_o   = [p for p in pairs if 0.4 <= p["jaccard"] < 0.6]
    med_o    = [p for p in pairs if threshold <= p["jaccard"] < 0.4]

    for p in critical:
        sample = ", ".join(p["shared_tokens"][:10])
        actions.append((
            "CRITICAL",
            "Token Overlap — Confusion Risk",
            f"{p['intent_a']}  ↔  {p['intent_b']}",
            f"Jaccard {p['jaccard']:.0%} — very high shared vocabulary ({p['shared_count']} tokens). "
            f"NLU will likely misclassify between these. Shared: {sample}",
            "Audit all phrases in both intents. Rephrase or remove shared vocabulary. "
            "Consider merging if these intents are semantically identical."
        ))
    for p in high_o:
        sample = ", ".join(p["shared_tokens"][:8])
        actions.append((
            "HIGH",
            "Token Overlap",
            f"{p['intent_a']}  ↔  {p['intent_b']}",
            f"Jaccard {p['jaccard']:.0%} ({p['shared_count']} shared tokens). Sample: {sample}",
            "Review and diversify training phrases. Add entity-based differentiation where possible."
        ))
    for p in med_o:
        sample = ", ".join(p["shared_tokens"][:6])
        actions.append((
            "MEDIUM",
            "Token Overlap",
            f"{p['intent_a']}  ↔  {p['intent_b']}",
            f"Jaccard {p['jaccard']:.0%} ({p['shared_count']} shared tokens). Sample: {sample}",
            "Monitor confusion in production. Consider adding distinguishing phrases."
        ))

    # WORD COUNT ISSUES
    for intent, s in sorted(stats.items()):
        if s["avg_wc"] < LOW_WORD_MIN:
            actions.append((
                "MEDIUM",
                "Phrase Length — Too Short",
                intent,
                f"Avg word count is {s['avg_wc']} — overly terse phrases reduce NLU generalisation.",
                "Add longer, more natural conversational phrases (aim for 4–10 words on average)."
            ))
        elif s["avg_wc"] > HIGH_WORD_MAX:
            actions.append((
                "LOW",
                "Phrase Length — Too Long",
                intent,
                f"Avg word count is {s['avg_wc']} — long phrases may overtrain on specific patterns.",
                "Supplement with shorter, more direct utterances. Keep a healthy mix of lengths."
            ))

    # TOP UBIQUITOUS TOKENS
    top_ubiq = [(t, d) for t, d in shared_token_map.items() if len(d) >= 8][:10]
    for token, intent_dict in top_ubiq:
        actions.append((
            "LOW",
            "Ubiquitous Token",
            f'"{token}"  (in {len(intent_dict)} intents)',
            f"Token '{token}' appears in {len(intent_dict)} different intents — it provides no discriminative signal.",
            "Consider whether this token can be replaced with more specific synonyms in some intents, "
            "or rely on surrounding context tokens to differentiate."
        ))

    # Priority sort order
    p_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    actions.sort(key=lambda x: p_order.get(x[0], 9))

    priority_colors = {
        "CRITICAL": ("FF4444", C_WHITE),
        "HIGH":     ("FF9933", C_WHITE),
        "MEDIUM":   ("FFD966", "7F6000"),
        "LOW":      ("E2EFDA", "375623"),
    }

    for i, (priority, category, subject, issue, action) in enumerate(actions):
        bg_color, fg_color = priority_colors.get(priority, (C_WHITE, "2C2C2C"))
        vals = [priority, category, subject, issue, action]
        write_data_row(ws, next_row, vals, stripe=False)

        p_cell = ws.cell(row=next_row, column=1)
        p_cell.fill = fill(bg_color)
        p_cell.font = Font(name="Arial", size=9, bold=True, color=fg_color)
        p_cell.alignment = centre()

        # Wrap text for issue + action columns
        for col in [4, 5]:
            ws.cell(row=next_row, column=col).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[next_row].height = 45

        next_row += 1

    # Summary counts
    next_row += 1
    counts = Counter(a[0] for a in actions)
    ws.cell(row=next_row, column=1, value="Summary").font = Font(
        name="Arial", bold=True, size=10)
    next_row += 1
    for level in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        ws.cell(row=next_row, column=1, value=level)
        ws.cell(row=next_row, column=2, value=counts.get(level, 0))
        bg, fg = priority_colors.get(level, (C_WHITE, "2C2C2C"))
        ws.cell(row=next_row, column=1).fill = fill(bg)
        ws.cell(row=next_row, column=1).font = Font(
            name="Arial", size=9, bold=True, color=fg)
        ws.cell(row=next_row, column=2).font = Font(name="Arial", size=9)
        next_row += 1

    set_col_widths(ws, [11, 30, 40, 60, 70])
    ws.freeze_panes = "A4"

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print(f"Loading {INPUT_CSV}...")
    intent_phrases = load_csv(INPUT_CSV)
    print(f"  {len(intent_phrases)} intents loaded.")

    print("Computing stats...")
    stats = compute_intent_stats(intent_phrases)

    print("Computing pairwise token overlap...")
    intents, pairs = compute_pairwise_overlap(stats)
    print(f"  {len(pairs)} pairs evaluated.")

    print("Building shared token map...")
    shared_token_map = compute_shared_token_map(stats)
    print(f"  {len(shared_token_map)} shared tokens found.")

    print("Building Excel report...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_summary(wb, stats, OVERLAP_THRESHOLD)
    build_overlap_matrix(wb, intents, stats)
    build_high_overlap(wb, pairs, OVERLAP_THRESHOLD)
    build_shared_tokens(wb, shared_token_map, TOP_N_SHARED)
    build_word_stats(wb, stats, intent_phrases)
    build_action_items(wb, stats, pairs, shared_token_map, OVERLAP_THRESHOLD)

    wb.save(OUTPUT_XLSX)
    print(f"\n✅  Report saved → {OUTPUT_XLSX}")

    # Console summary
    flagged_low  = sum(1 for s in stats.values() if s["phrase_count"] < LOW_PHRASE_MIN)
    high_overlap = sum(1 for p in pairs if p["jaccard"] >= OVERLAP_THRESHOLD)
    print(f"\n--- Quick Stats ---")
    print(f"Total intents:               {len(stats)}")
    print(f"Total phrases:               {sum(s['phrase_count'] for s in stats.values())}")
    print(f"Intents with < {LOW_PHRASE_MIN} phrases:   {flagged_low}")
    print(f"High overlap pairs (≥{OVERLAP_THRESHOLD}):  {high_overlap}")
    print(f"Unique shared tokens:        {len(shared_token_map)}")

if __name__ == "__main__":
    main()
